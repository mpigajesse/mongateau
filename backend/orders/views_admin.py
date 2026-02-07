"""
Admin-specific views for orders management
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Count, Sum, Q
from django.utils import timezone
from datetime import timedelta
from .models import Order, CustomCakeRequest
from .serializers import OrderSerializer, CustomCakeRequestSerializer, CustomCakeRequestCreateSerializer


class AdminOrderViewSet(viewsets.ModelViewSet):
    """
    Admin ViewSet for managing orders with advanced features
    """
    queryset = Order.objects.filter(is_deleted=False).order_by('-created_at')
    serializer_class = OrderSerializer

    def perform_destroy(self, instance):
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.save()

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """
        Get comprehensive order statistics for the dashboard
        """
        # Date ranges
        today = timezone.now().date()
        last_30_days = today - timedelta(days=30)
        last_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        last_month_end = today.replace(day=1) - timedelta(days=1)
        
        # Total orders
        total_orders = Order.objects.filter(is_deleted=False).count()
        
        # Orders by status
        pending_orders = Order.objects.filter(status='pending', is_deleted=False).count()
        confirmed_orders = Order.objects.filter(status='confirmed', is_deleted=False).count()
        delivered_orders = Order.objects.filter(status='delivered', is_deleted=False).count()
        cancelled_orders = Order.objects.filter(status='cancelled', is_deleted=False).count()
        
        # Revenue calculations
        total_revenue = Order.objects.filter(
            status='delivered',
            is_deleted=False
        ).aggregate(Sum('total_price'))['total_price__sum'] or 0
        
        # Revenue for last 30 days
        revenue_30_days = Order.objects.filter(
            status='delivered',
            created_at__gte=last_30_days,
            is_deleted=False
        ).aggregate(Sum('total_price'))['total_price__sum'] or 0
        
        # Revenue for last month
        revenue_last_month = Order.objects.filter(
            status='delivered',
            created_at__gte=last_month_start,
            created_at__lte=last_month_end,
            is_deleted=False
        ).aggregate(Sum('total_price'))['total_price__sum'] or 0
        
        # Calculate growth percentage
        revenue_growth = 0
        if revenue_last_month > 0:
            revenue_growth = ((revenue_30_days - revenue_last_month) / revenue_last_month) * 100
        
        # Orders growth
        orders_30_days = Order.objects.filter(created_at__gte=last_30_days, is_deleted=False).count()
        orders_last_month = Order.objects.filter(
            created_at__gte=last_month_start,
            created_at__lte=last_month_end,
            is_deleted=False
        ).count()
        
        orders_growth = 0
        if orders_last_month > 0:
            orders_growth = ((orders_30_days - orders_last_month) / orders_last_month) * 100
        
        # Most popular cake types
        popular_cakes = Order.objects.filter(is_deleted=False).values(
            'cake_type__name', 'cake_type__cake_type'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:5]
        
        # Recent orders
        recent_orders = Order.objects.filter(is_deleted=False).order_by('-created_at')[:5]
        
        # Custom requests count
        custom_requests_count = CustomCakeRequest.objects.count()
        pending_custom_requests = CustomCakeRequest.objects.filter(
            status='pending'
        ).count()
        
        return Response({
            'total_orders': total_orders,
            'pending_orders': pending_orders,
            'confirmed_orders': confirmed_orders,
            'delivered_orders': delivered_orders,
            'cancelled_orders': cancelled_orders,
            'total_revenue': float(total_revenue),
            'revenue_30_days': float(revenue_30_days),
            'revenue_growth': round(revenue_growth, 2),
            'orders_growth': round(orders_growth, 2),
            'popular_cakes': list(popular_cakes),
            'recent_orders': OrderSerializer(recent_orders, many=True).data,
            'custom_requests_count': custom_requests_count,
            'pending_custom_requests': pending_custom_requests,
        })
    
    @action(detail=False, methods=['get'])
    def revenue_chart(self, request):
        """
        Get revenue data for charts (last 12 months)
        """
        from django.db.models.functions import TruncMonth
        
        twelve_months_ago = timezone.now() - timedelta(days=365)
        
        revenue_by_month = Order.objects.filter(
            status='delivered',
            created_at__gte=twelve_months_ago,
            is_deleted=False
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            revenue=Sum('total_price'),
            count=Count('id')
        ).order_by('month')
        
        return Response(list(revenue_by_month))
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """
        Update order status
        """
        order = self.get_object()
        new_status = request.data.get('status')
        
        if new_status not in ['pending', 'confirmed', 'delivered', 'cancelled']:
            return Response(
                {'error': 'Invalid status'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        order.status = new_status
        order.save()
        
        return Response({
            'message': 'Status updated successfully',
            'order': OrderSerializer(order).data
        })
    
    @action(detail=False, methods=['post'])
    def bulk_update_status(self, request):
        """
        Update status for multiple orders
        """
        order_ids = request.data.get('order_ids', [])
        new_status = request.data.get('status')
        
        if not order_ids or not new_status:
            return Response(
                {'error': 'order_ids and status are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        updated_count = Order.objects.filter(
            id__in=order_ids
        ).update(status=new_status)
        
        return Response({
            'message': f'{updated_count} orders updated successfully',
            'updated_count': updated_count
        })
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export orders to Excel (XLSX)
        """
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Commandes"
        
        headers = [
            "N° Commande", "Client", "Téléphone", "Type de gâteau",
            "Date livraison", "Statut", "Prix total (FCFA)", "Créée le"
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        sheet.append(headers)
        for col_num, cell in enumerate(sheet[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            sheet.column_dimensions[get_column_letter(col_num)].width = 20
        
        for order in self.get_queryset():
            row = [
                order.order_number,
                order.customer_name,
                order.customer_phone,
                order.cake_type.name if order.cake_type else "N/A",
                order.delivery_date.strftime('%d/%m/%Y'),
                order.get_status_display(),
                float(order.total_price),
                order.created_at.strftime('%d/%m/%Y %H:%M')
            ]
            sheet.append(row)
        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
                cell.border = border
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="commandes.xlsx"'
        workbook.save(response)
        return response


class AdminCustomRequestViewSet(viewsets.ModelViewSet):
    """
    Admin ViewSet for managing custom cake requests
    """
    queryset = CustomCakeRequest.objects.filter(is_deleted=False).order_by('-created_at')
    serializer_class = CustomCakeRequestSerializer

    def perform_destroy(self, instance):
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.save()
    
    def get_serializer_class(self):
        if self.action == 'create':
            return CustomCakeRequestCreateSerializer
        return CustomCakeRequestSerializer
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """
        Update custom request status
        """
        custom_request = self.get_object()
        new_status = request.data.get('status')
        notes = request.data.get('notes', '')
        
        if new_status not in ['pending', 'reviewed', 'approved', 'rejected']:
            return Response(
                {'error': 'Invalid status'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        custom_request.status = new_status
        if notes:
            custom_request.admin_notes = notes
        custom_request.save()
        
        return Response({
            'message': 'Status updated successfully',
            'request': CustomCakeRequestSerializer(custom_request).data
        })
    
    @action(detail=True, methods=['post'])
    def send_quote(self, request, pk=None):
        """
        Send a quote to customer for custom request
        """
        custom_request = self.get_object()
        quote_amount = request.data.get('quote_amount')
        quote_details = request.data.get('quote_details', '')
        
        if not quote_amount:
            return Response(
                {'error': 'quote_amount is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update request with quote
        custom_request.estimated_budget = quote_amount
        custom_request.admin_notes = quote_details
        custom_request.status = 'reviewed'
        custom_request.save()
        
        # TODO: Send email/SMS to customer with quote
        
        return Response({
            'message': 'Quote sent successfully',
            'request': CustomCakeRequestSerializer(custom_request).data
        })
